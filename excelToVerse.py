# %% [markdown]
# # excel 파일 로드

# %%
from openpyxl import load_workbook
import re

load_wb  = load_workbook("dataTable.xlsm", data_only = True)
load_meta = load_wb["_meta_"]

parsePath = load_meta.cell(1, 2).value
closeConsole = load_meta.cell(2, 2).value

# %%
for sheetName in load_wb.sheetnames:
    if(sheetName == "_meta_" or sheetName == "!!readme!!"):
        continue
    
    print()
    print(sheetName)
    
    sheet = load_wb[sheetName] 
    for row in sheet.rows:
        for cell in row:
            print(cell.value, end=", ")
        print()

# %% [markdown]
# # excel 파일의 데이터를 파싱

# %%
class FieldItem:
    def __init__(self):
        self.field_name = ""
        self.type_name = ""
        self.values = list()
        self.row_values = dict()
        
    def __str__(self):
        s = f"[field_name: {self.field_name}, type_name={self.type_name}, values: ""{"
        
        for value in self.values:
            s += value + ", "
        s += "}"
            
        return s
    
def create_fieldItems(sheet):
    items = list()
    genField = False
    genType = False
    
    fieldCount = 0
    
    for row in sheet.rows:
        for cell in row:
            if(genField == False):
                item = FieldItem()
                item.field_name = str(cell.value)
                items.append(item)
            elif(genType == False):
                item = items[fieldCount]
                fieldCount += 1
                item.type_name = str(cell.value)
            else:
                item = items[fieldCount]
                if(fieldCount == 1):
                    item.key = str(cell.value)
                item.values.append(str(cell.value))
                fieldCount += 1
            
                
        if(genField == False):
            genField = True
        elif(genType == False):
            genType = True
        fieldCount = 0
    
    r_items = {}
    
    for item in items:
        r_items[item.field_name] = item
        
    return r_items
        

#test_field_items = create_fieldItems(load_wb["_weapon_"])
#print("keys: ")
#for item in test_field_items.keys():
#    print(item)
#print("\nvalues: ")
#for item in test_field_items.values():
#    print(item)
            
                
            

# %% [markdown]
# ## 데이터 유효성 검사

# %%
def make_valid_int_value(value):
    if(value.isdigit() == False):
        return "!!!유효하지_않은_int값형식!!!"
    
    index = value.find('.')
    
    if(index != -1):    
        return value[0:index - 1]
    
    return value



# %%

float_regex = re.compile("[0-9]")
def make_valid_float_value(value):
    if(value.count(".") > 1 or float_regex.match(value) == None):
        return "!!!유효하지_않은_float값형식!!!"
    
    index = value.find('.')
    
    if(index == -1):    
        value += ".0"
    
    return value



# %%
def make_valid_string_value(value):
    if(value.count("\"") > 0):
        return "!!!유효하지_않은_string값형식!!!"
    
    return f"\"{value}\""
        



# %%
def make_valid_logic_value(value):
    value = value.lower()
    if(value == "true" or value == "false"):
        return value
    
    return "!!!유효하지_않은_logic값형식!!!"

# %%
def make_valid_value(type, value):
    predicate = {
        "string": make_valid_string_value,
        "logic": make_valid_logic_value,
        "int": make_valid_int_value,
        "float": make_valid_float_value,
    }
    
    return predicate[type](value)

# %%
def make_default_value(type):
    predicate = {
        "string": "\"\"",
        "logic": "false",
        "int": "0",
        "float": "0.0",
    }
    
    return predicate[type]

# %%
print("==string==")
print(make_valid_string_value("123"))
print(make_valid_string_value("123.123321"))
print(make_valid_string_value("123.00.0"))
print(make_valid_string_value("ABCDE"))
print(make_valid_string_value("AB123CDE123"))
print(make_valid_string_value("A_=B123CDE12*3-"))
print(make_valid_string_value("AB12.3CDE123"))
print(make_valid_string_value("AB12.3CDE1.23"))

print("==int==")
print(make_valid_int_value("123"))
print(make_valid_int_value("123.123321"))
print(make_valid_int_value("123.00.0"))
print(make_valid_int_value("ABCDE"))
print(make_valid_int_value("AB123CDE123"))
print(make_valid_int_value("A_=B123CDE12*3-"))
print(make_valid_int_value("AB12.3CDE123"))
print(make_valid_int_value("AB12.3CDE1.23"))

print("==float==")
print(make_valid_float_value("123"))
print(make_valid_float_value("123.123321"))
print(make_valid_float_value("123.00.0"))
print(make_valid_float_value("ABCDE"))
print(make_valid_float_value("AB123CDE123"))
print(make_valid_float_value("A_=B123CDE12*3-"))
print(make_valid_float_value("AB12.3CDE123"))
print(make_valid_float_value("AB12.3CDE1.23"))

print("==logic==")
print(make_valid_logic_value("123"))
print(make_valid_logic_value("123.123321"))
print(make_valid_logic_value("123.00.0"))
print(make_valid_logic_value("ABCDE"))
print(make_valid_logic_value("AB123CDE123"))
print(make_valid_logic_value("A_=B123CDE12*3-"))
print(make_valid_logic_value("AB12.3CDE123"))
print(make_valid_logic_value("AB12.3CDE1.23"))
print(make_valid_logic_value("true"))
print(make_valid_logic_value("false"))
print(make_valid_logic_value("True"))
print(make_valid_logic_value("False"))
print(make_valid_logic_value("TRUE"))
print(make_valid_logic_value("FALSE"))

# %% [markdown]
# # verse 코드로 파싱

# %%
class SheetItem:
    def __init__(self, sheet, sheet_name):
        self.sheet = sheet
        self.sheet_name = sheet_name[1:len(sheet_name)-1]
        
        self.field_items = create_fieldItems(sheet)
        

# %%
def get_item_template(sheet_item):
    template = f"generated_{sheet_item.sheet_name}<public> := class<unique>():\n"
    indent = "    "
    
    field_template = indent + "var {field_name}<public>: {type_name} = {default_value}\n"
    
    for item in sheet_item.field_items.values():
        template += field_template.format(
            field_name = item.field_name,
            type_name = item.type_name,
            default_value = make_default_value(item.type_name)
        )
    
    return template

# %%
def get_constructor_item_template(sheet_item):
    template = f"generated_{sheet_item.sheet_name}_constructor<constructor>("
    indent = "    "
    count = 0
    
    arg_template = "Arg{number}: {type_name}"
    field_template = "{field_name} := Arg{number}"
    
    for item in sheet_item.field_items.values():
        template += arg_template.format(
            number = count,
            type_name = item.type_name
        )
        
        count += 1
        if(len(sheet_item.field_items.values()) != count):
            template += ", "
        else:
            template += f") := generated_{sheet_item.sheet_name}:\n"
    
    count = 0
    for item in sheet_item.field_items.values():
        template += indent + field_template.format(
            field_name = item.field_name,
            number = count
        ) + "\n"
        
        count += 1
    
    return template

# %%
def get_item_set_template(sheet_item):
    indent = "    "
    count = 0
    template = f"generated_{sheet_item.sheet_name}_set<public> := class<unique>():\n{indent}var Table<public>: [string]generated_{sheet_item.sheet_name} = map{{}}\n\n{indent}Initialize<public>():void=\n{indent}{indent}var Temp: generated_{sheet_item.sheet_name} = generated_{sheet_item.sheet_name}{{}}\n"
    
    item_template = "{indent}set Temp = {constructor_template}\n{indent}{indent}if(set Table[\"{field_name}\"] = Temp) {{}}"
    list_item = list(sheet_item.field_items.values())
    
    for index in range(0, len(list_item[0].values)):
        constructor_template = f"generated_{sheet_item.sheet_name}_constructor("
        key = ""
        for item in sheet_item.field_items.values():
            constructor_template += make_valid_value(item.type_name, item.values[index])
            if(count == 1):
                key = item.values[index]
            count += 1
            if(len(sheet_item.field_items.values()) != count):
                constructor_template += ", "
            else:
                constructor_template += ")"
        count = 0

        template += indent + item_template.format(
            field_name = key,
            sheet_name = sheet_item.sheet_name,
            constructor_template = constructor_template,
            class_name = f"generated_{sheet_item.sheet_name}",
            indent = indent
        ) + "\n"
    
    return template

# %%
def get_item_data_manager_function(sheet_item_list):
    template = ""
    template += "data_manager := class(base_data_manager):\n"
    template += "{fields}\n"
    template += "    Initialize<override>():void=\n"
    template += "{inits}\n"
    template += "{getters}\n"


    
    member_indent = "    "
    field_indent = "        "
    
    fields_output_template = ""
    getters_output_template = ""
    inits_output_template = ""
    
    fields_template = "{member_indent}var {sheet_name}<protected>: generated_{sheet_name}_set = generated_{sheet_name}_set{{}}\n"
    getters_template = "{member_indent}Get_{sheet_name}<public>(): generated_{sheet_name}_set=\n{field_indent}return {sheet_name}\n"
    getters_failure_template = "{member_indent}Get_F{sheet_name}<public>()<decides><transacts>: generated_{sheet_name}_set=\n{field_indent}return {sheet_name}\n"
    inits_template = "{field_indent}{sheet_name}.Initialize()\n"


    for sheet_item in sheet_item_list:
        fields_output_template += fields_template.format(
            member_indent = member_indent,
            sheet_name = sheet_item.sheet_name
        )
        
    for sheet_item in sheet_item_list:
        getters_output_template += getters_template.format(
            field_indent= field_indent,
            sheet_name = sheet_item.sheet_name,
            member_indent = member_indent
        ) + getters_failure_template.format(
            field_indent= field_indent,
            sheet_name = sheet_item.sheet_name,
            member_indent = member_indent
        )
        
    for sheet_item in sheet_item_list:
        inits_output_template += inits_template.format(
            sheet_name = sheet_item.sheet_name,
            field_indent= field_indent
        )
        
    return template.format(
        fields = fields_output_template,
        getters = getters_output_template,
        inits = inits_output_template
    )

# %%
#test_sheet_item = SheetItem(load_wb["_weapon_"], "_weapon_")
#print(get_item_template(test_sheet_item))
#print(get_constructor_item_template(test_sheet_item))
#print(get_item_set_template(test_sheet_item))
#print(get_item_injection_interface(test_sheet_item))
#print(get_item_data_manager_function([SheetItem(load_wb["_weapon_"], "_weapon_"), SheetItem(load_wb["_object_"], "_object_")]))

# %% [markdown]
# # 출력파일 생성

# %%
parsePath = load_meta.cell(1, 2).value
closeConsole = load_meta.cell(2, 2).value

s = ""

sheet_item_list = []
for sheet in load_wb.sheetnames:
    load_ws = None
    
    if(len(sheet) == 0):
        continue
    if(not (sheet[0] == '_' and sheet[-1] == '_')):
        continue
    if(sheet == "_meta_"):
        continue
    
    try:
        load_ws = load_wb[sheet]
    except:
        print("sheet load failed")
        pass
    
    try:
        print("-----------try to parse {name}-----------".format(name=sheet))
        print()
        p = SheetItem(load_ws, sheet)
        sheet_item_list.append(p)
        s += "\n#=============================================={name}==============================================\n".format(name = sheet)
        s += get_item_template(p) + "\n"
        s += get_constructor_item_template(p) + "\n"
        s += get_item_set_template(p) + "\n"
        s += "\n#==============================================\n#==============================================\n\n"
    except Exception as e:
        print(sheet + " sheet parse faield")
        print(e)
        
try:
    s += "\n#=============================================={{generated data_manager}}==============================================\n".format(name = sheet)
    s += get_item_data_manager_function(sheet_item_list)
except Exception as e:
        print(sheet + " sheet parse faield")
        print(e)
try:
    with open(parsePath+"\data.verse", 'w') as writer:
       writer.write(s + "\n")
except:
    print("file writing error")
        
if(not closeConsole):
    input("exit ?")


